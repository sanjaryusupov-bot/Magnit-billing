"""
Парсинг листов гугл-таблицы MCOS в плоские (tidy) таблицы для дашборда.

Лист "Статус по отгрузкам" — это визуальный отчёт с группировками:
    День (строка с датой, без заказа)
      Рейс N (строка-заголовок)
        строки заказов (дата, заказ, кол-во шт, магазин, регион, авто, сумма...)

Лист "Статус по заказам" — тоже с группировками по дням в колонке A:
    "01/07/2026 (заказы на ср)"  <- заголовок раздела (это и есть план. дата)
      строки заказов (№ заказа, ID магазина, магазин, кол-во шт, дата отгрузки план, город, статусы...)

============================================================================
ИЗМЕНЕНИЯ (сквозные тарифы, допсоглашение 2026):
============================================================================
1. DEFAULT_TARIFFS обновлён под новую консолидированную тарифную сетку:
   одиночные маршруты (как раньше) + сквозные маршруты (новое) — сквозные
   строки задаются тем же форматом, что и в листе "Тарифы Логистика" в
   Google-таблице: в колонке "Регион" перечислены ВСЕ города маршрута через
   запятую (например "Коканд, Наманган").
2. DEFAULT_REGION_TARIFF_MAP теперь пустой (identity) — по новому
   допсоглашению у каждого города своя ставка, старое сворачивание
   спутников в "Ташкент"/"Фергана" больше не действует.
3. Новое: build_tariff_indexes() — строит (одиночные, сквозные) индексы из
   DataFrame тарифов, ТОЧНО той же логикой сопоставления, что и в Apps
   Script calculateLogisticsRates (набор регионов рейса + тип авто).
4. add_route_economics() теперь:
     - помечает каждую строку колонками "Тип_тарифа" ("Сквозной"/"Обычный")
       и "Тариф_название" (человекочитаемое имя применённого тарифа) —
       чтобы удобно подсвечивать/фильтровать в отчёте;
     - для рейсов, распознанных как сквозные, НЕ разбивает Маршрут_ID по
       региону (в отличие от обычных рейсов) — весь сквозной рейс это один
       маршрут. Это чинит проблему "нулей": раньше Сумма_маршрута и
       Тариф_за_шт считались отдельно по каждому городу сквозного рейса,
       и все деньги (базовая ставка) доставались только тому городу, где
       была первая по рейсу точка — остальные города получали 0.
       Теперь Сумма_маршрута/Тариф_за_шт считаются на весь сквозной рейс
       целиком, и Сумма_распределенная размазывается по ВСЕМ строкам рейса
       пропорционально кол-ву штук — нулей быть не должно (кроме строк с
       Кол_во_шт = 0).
     - доплата "грузчик-экспедитор" теперь по умолчанию применяется КО
       ВСЕМ маршрутам (а не только LONG_HAUL_LOADER_FEE_REGIONS), т.к. в
       новом допсоглашении (п.4.5) явно сказано "тариф применяется по
       любому направлению, включая сквозные маршруты, и для любого типа
       транспортного средства". Старое поведение (только длинные плечи)
       можно вернуть флагом loader_fee_all_regions=False.
5. Новое: build_summary_df() — сводная по разрезам Регион / Авто / Рейс /
   Тариф, как просили: строится на уровне (Дата, Регион, Авто, Рейс,
   Тип_тарифа, Тариф_название).
6. build_billing_workbook() подсвечивает строки со Тип_тарифа="Сквозной"
   зелёной заливкой (тот же цвет, что и в файле тарифов) и добавляет
   колонки "Тип тарифа"/"Тариф" в реестр.
"""
import re
from datetime import datetime, date, timedelta
from io import BytesIO

import openpyxl
import pandas as pd

SHIPMENTS_SHEET = "Статус по отгрузкам"
ORDERS_SHEET = "Статус по заказам"
TARIFFS_SHEET = "тарифы"

_DATE_RE = re.compile(r"(\d{2})/(\d{2})/(\d{4})")

# Колонки (1-based), которые в исходных листах содержат дату и должны быть
# приведены к datetime, даже когда данные пришли как "сырые" числа (Google
# Sheets API отдаёт даты как порядковый номер дня, как в Excel).
_SHIPMENTS_DATE_COLS = {2}   # B: Дата отгрузки
_ORDERS_DATE_COLS = {6}      # F: Дата отгрузки план

_EXCEL_EPOCH = datetime(1899, 12, 30)

# Регион (точная точка доставки) -> Регион тариф (укрупнённая тарифная зона).
# ПО НОВОМУ ДОПСОГЛАШЕНИЮ (2026): у каждого города — своя согласованная
# ставка (см. раздел 2 доп.соглашения), поэтому сворачивать
# Янгиюль/Ангрен/Чирчик/Сырдарья/Коканд в "Ташкент"/"Фергана" больше не
# нужно. Карта оставлена пустой (identity) — Регион_тариф == Регион для
# всех городов. Если у вас остались специфичные исключения — добавьте их
# сюда явно.
DEFAULT_REGION_TARIFF_MAP: dict = {}

# Регионы (по Регион_тариф), для которых раньше ОГРАНИЧИВАЛСЯ пересчёт
# доплаты "грузчик-экспедитор" (см. loader_fee_all_regions в
# add_route_economics). Оставлено для обратной совместимости/отладки —
# по умолчанию больше НЕ используется, т.к. по новому допсоглашению (п.4.5)
# доплата грузчика применяется по любому направлению и типу ТС.
LONG_HAUL_LOADER_FEE_REGIONS = {"Самарканд", "Фергана", "Наманган", "Андижан", "Навои", "Бухара"}

# Порог кол-ва штук на маршрут, при достижении/превышении которого положена
# доплата "грузчик-экспедитор" (п.4.5 доп.соглашения).
LOADER_FEE_MIN_QTY = 3000
LOADER_FEE_AMOUNT = 440000

# ---------------------------------------------------------------------------
# Справочник тарифов по умолчанию — консолидированная сетка из
# Дополнительного соглашения 2026 (одиночные маршруты, раздел 2, + сквозные
# маршруты, раздел 3). Названия авто соответствуют тем, что реально стоят в
# выпадающем списке на листе "Статус по отгрузкам" (проверено по факту:
# "Газель", "Исузу 5 тонн", "10 тонн", "Лабо", "Евро фура" — БЕЗ "2 т"/"5 т"
# в тексте, как было в самом допсоглашении).
#
# Формат — 6 колонок, как раньше: № п/п, Регион, Авто, Маршрут, Базовая
# ставка за рейс, Стоимость доп. точки.
#
# СКВОЗНЫЕ строки (раздел 3 доп.соглашения) отличаются только тем, что в
# колонке "Регион" перечислены ВСЕ города маршрута через запятую — это
# единственный признак, по которому build_tariff_indexes() отличает
# сквозной тариф от одиночного (та же логика, что в Apps Script).
# ---------------------------------------------------------------------------
DEFAULT_TARIFFS = [
    # --- Одиночные маршруты ---------------------------------------------
    (1, "Ташкент", "Лабо с грузоподъемностью до 500кг. (9:00 - 18:00)",
     "Доставка заказов по г. Ташкент, Лабо 0,5 т (9:00-18:00)", 400000, 55000),
    (2, "Ташкент", "Лабо", "Доставка заказов по г. Ташкент, Лабо 0,5 т", 210000, 55000),
    (3, "Ташкент", "Газель", "Доставка заказов по г. Ташкент, Газель", 450000, 165000),
    (4, "Ташкент", "Исузу 5 тонн", "Доставка заказов по г. Ташкент, Исузу 5 тонн", 935000, 165000),
    (5, "Ташкент", "10 тонн", "Доставка заказов по г. Ташкент, 10 тонн", 1430000, 165000),
    (6, "Ташкент", "Евро фура", "Доставка заказов по г. Ташкент, Евро фура", 1760000, 220000),
    (7, "Чирчик", "Газель", "Ташкент - Чирчик", 890000, 165000),
    (8, "Чирчик", "Исузу 5 тонн", "Ташкент - Чирчик", 1150000, 165000),
    (9, "Янгиюль", "Газель", "Ташкент - Янгиюль", 1020000, 165000),
    (10, "Янгиюль", "Исузу 5 тонн", "Ташкент - Янгиюль", 1280000, 165000),
    (11, "Нурафшон", "Газель", "Ташкент - Нурафшон", 1400000, 165000),
    (12, "Нурафшон", "Исузу 5 тонн", "Ташкент - Нурафшон", 1660000, 165000),
    (13, "Сырдарья", "Газель", "Ташкент - Сырдарья", 1400000, 165000),
    (14, "Сырдарья", "Исузу 5 тонн", "Ташкент - Сырдарья", 1660000, 165000),
    (15, "Ангрен", "Газель", "Ташкент - Ангрен", 1790000, 165000),
    (16, "Ангрен", "Исузу 5 тонн", "Ташкент - Ангрен", 1910000, 165000),
    (17, "Ангрен (транзит)", "Исузу 5 тонн",
     "Ташкент - Ангрен (транзитом в Ферганскую долину) — ДОПУЩЕНИЕ: применяется только когда "
     "Ангрен идёт транзитом в составе рейса в Ферганскую долину", 380000, 165000),
    (18, "Коканд", "Газель", "Ташкент - Коканд", 2550000, 165000),
    (19, "Коканд", "Исузу 5 тонн", "Ташкент - Коканд", 3000000, 165000),
    (20, "Наманган", "Газель", "Ташкент - Наманган", 2680000, 165000),
    (21, "Наманган", "Исузу 5 тонн", "Ташкент - Наманган", 2970000, 165000),
    (22, "Наманган", "10 тонн", "Ташкент - Наманган", 3960000, 220000),
    (23, "Фергана", "Газель", "Ташкент - Фергана", 2300000, 165000),
    (24, "Фергана", "Исузу 5 тонн", "Ташкент - Фергана", 2900000, 165000),
    (25, "Фергана", "10 тонн", "Ташкент - Фергана", 3900000, 220000),
    (26, "Андижан", "Газель", "Ташкент - Андижан", 2600000, 165000),
    (27, "Андижан", "Исузу 5 тонн", "Ташкент - Андижан", 3300000, 220000),
    # (28) Андижан, 10 тонн — тариф НЕ согласован (п.4.7 доп.соглашения), намеренно пропущен.
    (29, "Самарканд", "Газель", "Ташкент - Самарканд", 1900000, 165000),
    (30, "Самарканд", "Исузу 5 тонн", "Ташкент - Самарканд", 2420000, 165000),
    (31, "Самарканд", "10 тонн", "Ташкент - Самарканд", 2970000, 220000),
    (32, "Навои", "Газель", "Ташкент - Навои", 2700000, 220000),
    (33, "Навои", "Исузу 5 тонн", "Ташкент - Навои", 4400000, 220000),
    (34, "Навои", "10 тонн", "Ташкент - Навои", 5900000, 220000),
    (35, "Бухара", "Газель", "Ташкент - Бухара", 3950000, 220000),
    (36, "Бухара", "Исузу 5 тонн", "Ташкент - Бухара", 4650000, 220000),
    (37, "Бухара", "10 тонн", "Ташкент - Бухара", 6500000, 220000),
    # --- Сквозные (комбинированные) маршруты — раздел 3 доп.соглашения ---
    (38, "Коканд, Наманган", "Исузу 5 тонн", "Ташкент - Коканд - Наманган (сквозной)", 5000000, 165000),
    (39, "Коканд, Наманган", "Газель", "Ташкент - Коканд - Наманган (сквозной)", 3570000, 165000),
    (40, "Навои, Бухара", "Исузу 5 тонн", "Ташкент - Навои - Бухара (сквозной)", 7140000, 220000),
    (41, "Навои, Бухара", "Газель", "Ташкент - Навои - Бухара (сквозной)", 5360000, 220000),
    (42, "Самарканд, Навои, Бухара", "Исузу 5 тонн",
     "Ташкент - Самарканд - Навои - Бухара (сквозной)", 8040000, 220000),
    (43, "Коканд, Наманган, Андижан", "Исузу 5 тонн",
     "Ташкент - Коканд - Наманган - Андижан (сквозной)", 8040000, 220000),
    (44, "Нурафшон, Сырдарья", "Исузу 5 тонн", "Ташкент - Нурафшон - Сырдарья (сквозной)", 3210000, 165000),
    # --- Служебное -------------------------------------------------------
    (45, None, None, "Дополнительно грузчик-экспедитор в машину (при кол-ве шт >= 3000)", 440000, None),
]

_TARIFF_COLS = ["№ п/п", "Регион", "Авто", "Маршрут", "Базовая ставка за рейс", "Стоимость дополнительной точки"]


def default_tariffs_df() -> pd.DataFrame:
    return pd.DataFrame(DEFAULT_TARIFFS, columns=_TARIFF_COLS)


def _serial_to_datetime(v):
    """Преобразует порядковый номер даты (Google Sheets/Excel) в datetime."""
    try:
        return _EXCEL_EPOCH + timedelta(days=float(v))
    except (TypeError, ValueError):
        return None


class _Cell:
    __slots__ = ("value",)

    def __init__(self, value):
        self.value = value


class ListWorksheet:
    """
    Адаптер: превращает список списков (сырые значения, например от gspread)
    в объект с интерфейсом openpyxl Worksheet (.cell(row, column).value,
    .max_row), чтобы parse_shipments/parse_orders_status работали одинаково
    и с .xlsx, и с данными из Google Sheets API.
    """

    def __init__(self, rows, date_columns=None):
        self.rows = rows
        self.date_columns = date_columns or set()
        self.max_row = len(rows)

    def cell(self, row, column):
        r, c = row - 1, column - 1
        if r < 0 or r >= len(self.rows):
            return _Cell(None)
        row_data = self.rows[r]
        if c < 0 or c >= len(row_data):
            return _Cell(None)
        v = row_data[c]
        if v == "":
            v = None
        if v is not None and column in self.date_columns and isinstance(v, (int, float)) and not isinstance(v, bool):
            dt = _serial_to_datetime(v)
            if dt is not None:
                v = dt
        return _Cell(v)


def _is_error(v):
    if v is None:
        return True
    if isinstance(v, str) and v.strip().startswith("#"):
        return True
    return False


def _num(v):
    if _is_error(v):
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _clean_str(v):
    if _is_error(v):
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def parse_shipments(ws) -> pd.DataFrame:
    """Лист 'Статус по отгрузкам' -> tidy DataFrame, 1 строка = 1 заказ в рейсе."""
    rows = []
    current_day = None
    current_reys = None
    for r in range(2, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value  # Дата отгрузки
        c = ws.cell(row=r, column=3).value  # Заказ
        if b is None and c is None:
            continue
        if isinstance(b, str) and b.strip().lower().startswith("рейс"):
            current_reys = b.strip()
            continue
        if isinstance(b, datetime) and c is None:
            current_day = b.date()
            current_reys = None
            continue
        if c is not None:
            day_val = b.date() if isinstance(b, datetime) else current_day
            rows.append(
                {
                    "Дата": day_val,
                    "Рейс": current_reys,
                    "Заказ": str(c).strip(),
                    "Кол_во_шт": _num(ws.cell(row=r, column=4).value),
                    "ID_магазина": _clean_str(ws.cell(row=r, column=5).value),
                    "Магазин": _clean_str(ws.cell(row=r, column=6).value),
                    "Адрес": _clean_str(ws.cell(row=r, column=7).value),
                    "Координаты_1": _clean_str(ws.cell(row=r, column=8).value),
                    "Координаты_2": _clean_str(ws.cell(row=r, column=9).value),
                    "Регион": _clean_str(ws.cell(row=r, column=10).value),
                    "Транспорт": _clean_str(ws.cell(row=r, column=11).value),
                    "Точка_1": _num(ws.cell(row=r, column=12).value),
                    "Точка_2_и_далее": _num(ws.cell(row=r, column=13).value),
                    "Грузчик_экспедитор": _num(ws.cell(row=r, column=14).value),
                    "Итого_сумма": _num(ws.cell(row=r, column=15).value),
                    "Подрядчик": _clean_str(ws.cell(row=r, column=16).value),
                }
            )
    df = pd.DataFrame(rows)
    if not df.empty:
        df["Дата"] = pd.to_datetime(df["Дата"])
    return df


def parse_orders_status(ws) -> pd.DataFrame:
    """Лист 'Статус по заказам' -> tidy DataFrame с плановой датой отгрузки."""
    rows = []
    current_section_date = None
    for r in range(2, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if a is None:
            continue
        a_str = str(a).strip()
        if not a_str.startswith("1M-"):
            m = _DATE_RE.search(a_str)
            if m:
                d, mo, y = m.groups()
                try:
                    current_section_date = date(int(y), int(mo), int(d))
                except ValueError:
                    pass
            continue
        plan_cell = ws.cell(row=r, column=6).value  # Дата отгрузки план
        plan_date = plan_cell.date() if isinstance(plan_cell, datetime) else current_section_date
        rows.append(
            {
                "Заказ": a_str,
                "ID_магазина": _clean_str(ws.cell(row=r, column=2).value),
                "Магазин": _clean_str(ws.cell(row=r, column=3).value),
                "Адрес_магазина": _clean_str(ws.cell(row=r, column=4).value),
                "Кол_во_шт_заказ": _num(ws.cell(row=r, column=5).value),
                "Дата_план": plan_date,
                "Город": _clean_str(ws.cell(row=r, column=7).value),
                "Статус_сборки": _clean_str(ws.cell(row=r, column=8).value),
                "Статус_WMS": _clean_str(ws.cell(row=r, column=9).value),
                "Статус_отгрузки_на_хаб": _clean_str(ws.cell(row=r, column=10).value),
            }
        )
    df = pd.DataFrame(rows)
    if not df.empty:
        df["Дата_план"] = pd.to_datetime(df["Дата_план"])
    return df


# ---------------------------------------------------------------------------
# НОВОЕ: индекс тарифов (одиночные / сквозные) — та же логика сопоставления,
# что и в Apps Script calculateLogisticsRates.
# ---------------------------------------------------------------------------
def _normalize_regions(regions_str: str):
    return sorted({s.strip().lower() for s in str(regions_str).split(",") if s.strip()})


def build_tariff_indexes(tariffs_df: pd.DataFrame = None):
    """
    Возвращает (single_tariffs, through_tariffs):
      single_tariffs:  {(регион_lower, авто_lower): {"base":.., "extra":.., "name":..}}
      through_tariffs: {(tuple(отсортированные_регионы_lower), авто_lower): {...}}

    Строка считается сквозной, если в колонке "Регион" есть запятая
    (несколько городов через запятую) — точно так же, как определяет Apps
    Script при чтении листа "Тарифы Логистика".
    """
    if tariffs_df is None or tariffs_df.empty:
        tariffs_df = default_tariffs_df()

    single, through = {}, {}
    for _, row in tariffs_df.iterrows():
        region_raw = row.get("Регион")
        car_raw = row.get("Авто")
        if region_raw is None or car_raw is None:
            continue
        region_raw = str(region_raw).strip()
        car = str(car_raw).strip().lower()
        if not region_raw or not car:
            continue

        base = _num(row.get("Базовая ставка за рейс"))
        extra = _num(row.get("Стоимость дополнительной точки"))
        name = _clean_str(row.get("Маршрут")) or region_raw

        if "," in region_raw:
            regions = tuple(_normalize_regions(region_raw))
            through[(regions, car)] = {"base": base, "extra": extra, "name": name, "regions": regions}
        else:
            single[(region_raw.lower(), car)] = {"base": base, "extra": extra, "name": name}

    return single, through


def _match_through_tariff(regions: set, car: str, through_tariffs: dict):
    """Ищет сквозной тариф под набор регионов рейса + авто (порядок не важен)."""
    if not regions or len(regions) < 2 or not car:
        return None
    key = (tuple(sorted(r.lower() for r in regions if r)), car.lower())
    return through_tariffs.get(key)


def add_route_economics(
    ship_df: pd.DataFrame,
    region_tariff_map: dict = None,
    tariffs_df: pd.DataFrame = None,
    loader_fee_all_regions: bool = True,
) -> pd.DataFrame:
    """
    Пересчитывает стоимость доставки по логике заказчика и добавляет поля,
    нужные для биллинг-отчёта.

    В исходнике "Итого сумма" — это тариф за точку маршрута (магазин), который
    проставлен только один раз на уникальный магазин в рейсе (повторные заказы
    в тот же магазин в тот же рейс получают 0, доставка уже учтена). Из-за
    этого при группировке "по магазину" суммы получаются неровными — то 0, то
    полный тариф. Для СКВОЗНЫХ рейсов дополнительно вся базовая ставка рейса
    целиком проставлена только в ПЕРВОЙ по рейсу строке (см. Apps Script) —
    остальные города того же сквозного рейса могут вообще не иметь base в
    своей "родной" строке.

    Правильный расчёт (распределение):
        1. Определяем для каждого исходного блока (Дата, Рейс) — сквозной он
           или нет: смотрим набор уникальных "Регион" внутри блока + авто
           (первое непустое значение в блоке) и сверяем с through_tariffs
           (build_tariff_indexes). Тип сохраняется в колонке "Тип_тарифа"
           ("Сквозной"/"Обычный"), название применённого тарифа — в
           "Тариф_название".
        2. Маршрут_ID — единица группировки для распределения:
             - Сквозной блок -> ВЕСЬ блок это один маршрут (не разбиваем по
               региону!) — иначе города, не получившие базовую ставку в
               своей строке, останутся с нулевой распределённой суммой.
             - Обычный блок -> как раньше: если внутри блока встретились
               заказы из разных Регион_тариф, каждая зона — отдельный
               маршрут (Маршрут_ID = номер первого заказа этой зоны внутри
               блока).
        3. Сумма_маршрута = сумма "Итого сумма" по всем строкам маршрута.
           Кол_во_шт_маршрута = сумма "Кол-во шт" по всем строкам маршрута.
           Тариф_за_шт = Сумма_маршрута / Кол_во_шт_маршрута.
           Сумма_распределенная (для каждой строки) = Кол-во_шт этой строки
           * Тариф_за_шт — стоимость доставки размазана по всем штукам
           маршрута пропорционально, включая все города сквозного рейса.

    Дополнительно (для биллинг-реестра) — ДВА варианта идентификатора
    маршрута:
        - Маршрут_ID_первый_заказ — номер ПЕРВОГО заказа, встреченного в
          блоке (Дата, Рейс) исходника, без какого-либо разделения. Идёт в
          колонку "Рейс" биллинг-отчёта, отражает исходную разбивку
          "Рейс N" как есть.
        - Маршрут_ID — единица группировки для расчёта (см. выше).
        - Регион_тариф — регион, нормализованный к тарифной зоне (по
          умолчанию = сам регион, см. DEFAULT_REGION_TARIFF_MAP).

    Доплата "грузчик-экспедитор" (LOADER_FEE_AMOUNT = 440 000):
        - Считается по маршруту (Дата, Маршрут_ID) — суммарное кол-во штук.
        - По умолчанию (loader_fee_all_regions=True) применяется КО ВСЕМ
          маршрутам — по новому допсоглашению (п.4.5) доплата действует "по
          любому направлению, включая сквозные маршруты, и для любого типа
          транспортного средства". Передайте loader_fee_all_regions=False,
          чтобы вернуть старое поведение (только LONG_HAUL_LOADER_FEE_REGIONS).
        - "Грузчик экспедитор" пересчитывается с нуля (любое ошибочно
          проставленное в исходнике значение убирается): если суммарное
          кол-во штук по маршруту >= 3000, ровно одной (якорной) строке
          маршрута проставляется 440 000, у всех остальных строк маршрута —
          0. "Итого сумма" пересчитывается как "1 точка" + "2 точка и
          далее" + "Грузчик экспедитор".

    Группировка для расчёта тарифа за штуку — по (Дата, Маршрут_ID).
    """
    if ship_df.empty:
        return ship_df

    region_map = region_tariff_map if region_tariff_map is not None else DEFAULT_REGION_TARIFF_MAP
    single_tariffs, through_tariffs = build_tariff_indexes(tariffs_df)

    df = ship_df.copy()
    df["Регион_тариф"] = df["Регион"].map(lambda r: region_map.get(r, r) if r is not None else r)

    # Маршрут_ID_первый_заказ: наивный вариант — просто номер первого заказа
    # в блоке (Дата, Рейс), без разделения. Идёт в колонку "Рейс".
    df["Маршрут_ID_первый_заказ"] = df.groupby(
        ["Дата", "Рейс"], dropna=False, sort=False
    )["Заказ"].transform("first")

    # --- Определяем тип тарифа (Сквозной/Обычный) на уровне блока (Дата, Рейс) ---
    def _block_tariff_type(g: pd.DataFrame):
        regions = {r for r in g["Регион"].tolist() if r}
        car = None
        for c in g["Транспорт"].tolist():
            if c:
                car = c
                break
        match = _match_through_tariff(regions, car or "", through_tariffs)
        if match:
            return "Сквозной", match["name"]
        return "Обычный", None

    tip_tarifa = pd.Series(index=df.index, dtype=object)
    tariff_name = pd.Series(index=df.index, dtype=object)
    block_is_through = {}  # (Дата, Рейс) -> bool

    for key, idx in df.groupby(["Дата", "Рейс"], dropna=False, sort=False).groups.items():
        tip, name = _block_tariff_type(df.loc[idx])
        tip_tarifa.loc[idx] = tip
        block_is_through[key] = (tip == "Сквозной")
        if tip == "Сквозной":
            tariff_name.loc[idx] = name
        else:
            # для обычных — имя тарифа проставим ниже, по региону каждой строки
            tariff_name.loc[idx] = None

    df["Тип_тарифа"] = tip_tarifa

    # Имя тарифа для обычных строк — по одиночному тарифу (Регион_тариф + Авто),
    # если он есть в справочнике; иначе просто "{Регион_тариф}".
    def _single_tariff_name(row):
        if row["Тип_тарифа"] == "Сквозной":
            return tariff_name.loc[row.name]
        car = (row["Транспорт"] or "").strip().lower()
        reg = (row["Регион_тариф"] or "").strip().lower()
        m = single_tariffs.get((reg, car))
        if m:
            return m["name"]
        return row["Регион_тариф"]

    df["Тариф_название"] = df.apply(_single_tariff_name, axis=1)

    # --- Маршрут_ID: сквозной блок = 1 маршрут целиком; обычный блок = разбивка по региону ---
    def _assign_route_ids(g: pd.DataFrame, is_through: bool) -> pd.Series:
        if is_through:
            anchor = g["Заказ"].iloc[0]
            return pd.Series(anchor, index=g.index)
        orders = g["Заказ"].tolist()
        regions = g["Регион_тариф"].tolist()
        region_to_anchor = {}
        result = []
        for order, region in zip(orders, regions):
            anchor = region_to_anchor.setdefault(region, order)
            result.append(anchor)
        return pd.Series(result, index=g.index)

    marshrut_id = pd.Series(index=df.index, dtype=object)
    for key, idx in df.groupby(["Дата", "Рейс"], dropna=False, sort=False).groups.items():
        marshrut_id.loc[idx] = _assign_route_ids(df.loc[idx], block_is_through.get(key, False))
    df["Маршрут_ID"] = marshrut_id

    # --- Доплата "грузчик-экспедитор" ---------------------------------
    grp_cols = ["Дата", "Маршрут_ID"]
    qty_per_route = df.groupby(grp_cols, dropna=False)["Кол_во_шт"].transform("sum")
    is_anchor_row = df["Заказ"] == df["Маршрут_ID"]

    if loader_fee_all_regions:
        recalc_mask = pd.Series(True, index=df.index)
    else:
        recalc_mask = df["Регион_тариф"].isin(LONG_HAUL_LOADER_FEE_REGIONS)

    df.loc[recalc_mask, "Грузчик_экспедитор"] = 0.0
    eligible = recalc_mask & is_anchor_row & (qty_per_route >= LOADER_FEE_MIN_QTY)
    df.loc[eligible, "Грузчик_экспедитор"] = LOADER_FEE_AMOUNT

    df.loc[recalc_mask, "Итого_сумма"] = (
        df.loc[recalc_mask, "Точка_1"]
        + df.loc[recalc_mask, "Точка_2_и_далее"]
        + df.loc[recalc_mask, "Грузчик_экспедитор"]
    )
    # --------------------------------------------------------------------

    route_totals = (
        df.groupby(grp_cols, dropna=False)
        .agg(Сумма_маршрута=("Итого_сумма", "sum"), Кол_во_шт_маршрута=("Кол_во_шт", "sum"))
        .reset_index()
    )
    route_totals["Тариф_за_шт"] = route_totals.apply(
        lambda r: (r["Сумма_маршрута"] / r["Кол_во_шт_маршрута"]) if r["Кол_во_шт_маршрута"] else 0.0,
        axis=1,
    )
    df = df.merge(route_totals, on=grp_cols, how="left")
    df["Сумма_распределенная"] = df["Кол_во_шт"] * df["Тариф_за_шт"]

    return df


def build_summary_df(ship_df: pd.DataFrame) -> pd.DataFrame:
    """
    Сводная: разрезы Регион / Авто / Рейс / Тариф, как просили.

    Группировка — (Дата, Регион_тариф, Транспорт, Маршрут_ID_первый_заказ,
    Тип_тарифа, Тариф_название). Для сквозных рейсов внутри одного "Рейс"
    может быть несколько строк сводной (по одной на каждый город маршрута)
    — все они будут иметь одинаковые "Рейс"/"Тип тарифа"/"Тариф", так что их
    легко распознать и, при необходимости, свернуть отдельным пивотом.

    Требует, чтобы ship_df уже прошёл через add_route_economics().
    """
    if ship_df.empty:
        return pd.DataFrame()

    required = {"Регион_тариф", "Транспорт", "Маршрут_ID_первый_заказ", "Тип_тарифа", "Тариф_название"}
    missing = required - set(ship_df.columns)
    if missing:
        raise ValueError(
            f"build_summary_df: в ship_df нет колонок {missing} — сначала вызовите add_route_economics()"
        )

    grp_cols = ["Дата", "Регион_тариф", "Транспорт", "Маршрут_ID_первый_заказ", "Тип_тарифа", "Тариф_название"]
    summary = (
        ship_df.groupby(grp_cols, dropna=False)
        .agg(
            Кол_во_заказов=("Заказ", "nunique"),
            Кол_во_магазинов=("Адрес", "nunique"),
            Кол_во_шт=("Кол_во_шт", "sum"),
            Сумма_итого=("Итого_сумма", "sum"),
            Сумма_распределенная=("Сумма_распределенная", "sum"),
        )
        .reset_index()
        .rename(
            columns={
                "Регион_тариф": "Регион",
                "Транспорт": "Авто",
                "Маршрут_ID_первый_заказ": "Рейс",
                "Тариф_название": "Тариф",
            }
        )
        .sort_values(["Дата", "Рейс", "Регион"])
        .reset_index(drop=True)
    )
    return summary


def parse_tariffs(ws) -> pd.DataFrame:
    """Лист 'тарифы' (справочник ставок) -> DataFrame с теми же заголовками."""
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 7)]
        if all(v is None for v in vals):
            continue
        rows.append(vals)
    if not rows:
        return pd.DataFrame(columns=_TARIFF_COLS)
    return pd.DataFrame(rows, columns=_TARIFF_COLS)


def parse_workbook_bytes(data: bytes):
    """Возвращает (ship_df, orders_df) из байтов .xlsx файла."""
    wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
    ship_df = parse_shipments(wb[SHIPMENTS_SHEET])
    orders_df = parse_orders_status(wb[ORDERS_SHEET])
    return ship_df, orders_df


_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets.readonly",
    "https://www.googleapis.com/auth/drive.readonly",
]


def load_from_gsheet_service_account(sheet_id: str, credentials_info: dict):
    """
    Возвращает (ship_df, orders_df, tariffs_df), читая таблицу через Google
    Sheets API с помощью сервисного аккаунта. Таблицу не нужно открывать по
    ссылке — достаточно выдать доступ на чтение самому сервисному аккаунту
    (credentials_info['client_email']).

    Лист "тарифы" — необязательный: если такого листа в таблице нет,
    возвращается пустой DataFrame (в приложении в этом случае используется
    встроенный справочник по умолчанию, см. default_tariffs_df()).
    """
    import gspread
    from google.oauth2.service_account import Credentials

    creds = Credentials.from_service_account_info(dict(credentials_info), scopes=_SCOPES)
    client = gspread.authorize(creds)
    sh = client.open_by_key(sheet_id)

    ship_ws = sh.worksheet(SHIPMENTS_SHEET)
    orders_ws = sh.worksheet(ORDERS_SHEET)

    ship_rows = ship_ws.get_values(value_render_option="UNFORMATTED_VALUE")
    orders_rows = orders_ws.get_values(value_render_option="UNFORMATTED_VALUE")

    ship_adapter = ListWorksheet(ship_rows, date_columns=_SHIPMENTS_DATE_COLS)
    orders_adapter = ListWorksheet(orders_rows, date_columns=_ORDERS_DATE_COLS)

    ship_df = parse_shipments(ship_adapter)
    orders_df = parse_orders_status(orders_adapter)

    tariffs_df = pd.DataFrame(columns=_TARIFF_COLS)
    try:
        tariffs_ws = sh.worksheet(TARIFFS_SHEET)
        tariffs_rows = tariffs_ws.get_values(value_render_option="UNFORMATTED_VALUE")
        tariffs_adapter = ListWorksheet(tariffs_rows)
        tariffs_df = parse_tariffs(tariffs_adapter)
    except Exception:
        pass  # листа "тарифы" нет в таблице — это нормально, используем дефолт

    return ship_df, orders_df, tariffs_df


def build_sla(ship_df: pd.DataFrame, orders_df: pd.DataFrame) -> pd.DataFrame:
    """Джойн план (Статус по заказам) vs факт (Статус по отгрузкам) по номеру заказа 1M-."""
    if ship_df.empty or orders_df.empty:
        return pd.DataFrame()

    fact = (
        ship_df[ship_df["Заказ"].str.startswith("1M-")]
        .groupby("Заказ", as_index=False)
        .agg(Дата_факт=("Дата", "min"))
    )
    plan = orders_df[orders_df["Заказ"].str.startswith("1M-")][
        ["Заказ", "Магазин", "Город", "Дата_план", "Статус_отгрузки_на_хаб"]
    ]
    sla = plan.merge(fact, on="Заказ", how="left")
    sla["Дельта_дней"] = (sla["Дата_факт"] - sla["Дата_план"]).dt.days
    sla["SLA_статус"] = sla.apply(_sla_flag, axis=1)
    return sla


def _sla_flag(row):
    if pd.isna(row["Дата_план"]):
        return "Нет плана"
    if pd.isna(row["Дата_факт"]):
        return "Не отгружен"
    if row["Дельта_дней"] <= 0:
        return "В срок"
    return "Просрочка"


def build_billing_workbook(ship_df: pd.DataFrame) -> bytes:
    """
    Собирает Excel-файл с листами "Реестр" и "Сводная":

    "Реестр" — построчно по каждому заказу, колонки как раньше + новые
    "Тип тарифа" и "Тариф" — чтобы визуально (заливка) и через фильтр было
    видно, что посчитано по сквозному тарифу, а что по обычной сетке.
    Строки со Тип тарифа = "Сквозной" подсвечены зелёным (тот же цвет, что
    и в файле "Тарифы_новые.xlsx").

    "Сводная" — build_summary_df() (разрезы Регион / Авто / Рейс / Тариф).

    Оформлен как настоящая Excel-таблица (с автофильтром) — источник
    данных, из которого в Excel за 2 клика строится живая сводная таблица.

    ship_df должен быть уже отфильтрован (период, только 1M- и т.п.) и
    содержать колонки, которые добавляет add_route_economics().
    """
    from io import BytesIO as _BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = Workbook()
    DATE_FMT = "YYYY-MM-DD"
    THROUGH_FILL = PatternFill("solid", fgColor="E2EFDA")  # зелёный — сквозные строки
    HEADER_FONT = Font(bold=True)

    def _autosize(ws, ncols, min_width=10, max_width=45):
        for c in range(1, ncols + 1):
            letter = get_column_letter(c)
            max_len = min_width
            for row in ws.iter_rows(min_col=c, max_col=c):
                v = row[0].value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[letter].width = min(max_len + 2, max_width)

    # ================= Лист "Реестр" =================
    ws3 = wb.active
    ws3.title = "Реестр"
    registry_cols = [
        "Дата отгрузки", "Заказ", "Рейс", "Рейс 2", "Кол-во штук", "ID магазина", "Магазин", "Адрес",
        "Координаты", "Координаты2", "Регион", "Регион тариф", "Авто",
        "Тип тарифа", "Тариф",
        "1 точка", "2 точка и далее", "Грузчик экспедитор", "Итого сумма", "Сумма распределенная",
    ]
    ws3.append(registry_cols)
    for c in range(1, len(registry_cols) + 1):
        ws3.cell(row=1, column=c).font = HEADER_FONT

    through_col_idx = registry_cols.index("Тип тарифа") + 1

    n_data_rows = 0
    if not ship_df.empty:
        reg = ship_df.sort_values(["Дата", "Маршрут_ID", "Заказ"])
        for _, row in reg.iterrows():
            ws3.append(
                [
                    row["Дата"].date(),
                    row["Заказ"],
                    row["Маршрут_ID_первый_заказ"],
                    row["Маршрут_ID"],
                    row["Кол_во_шт"],
                    row["ID_магазина"],
                    row["Магазин"],
                    row["Адрес"],
                    row["Координаты_1"],
                    row["Координаты_2"],
                    row["Регион"],
                    row["Регион_тариф"],
                    row["Транспорт"],
                    row["Тип_тарифа"],
                    row["Тариф_название"],
                    row["Точка_1"],
                    row["Точка_2_и_далее"],
                    row["Грузчик_экспедитор"],
                    row["Итого_сумма"],
                    row["Сумма_распределенная"],
                ]
            )
            n_data_rows += 1
            if row["Тип_тарифа"] == "Сквозной":
                for c in range(1, len(registry_cols) + 1):
                    ws3.cell(row=ws3.max_row, column=c).fill = THROUGH_FILL

    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=1, max_col=1):
        row[0].number_format = DATE_FMT
    _autosize(ws3, len(registry_cols), max_width=40)

    if n_data_rows > 0:
        last_col_letter = get_column_letter(len(registry_cols))
        table_ref = f"A1:{last_col_letter}{n_data_rows + 1}"
        tbl = Table(displayName="ТаблицаРеестр", ref=table_ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False,
            showLastColumn=False, showColumnStripes=False,
        )
        ws3.add_table(tbl)

    # ================= Лист "Сводная" =================
    summary = build_summary_df(ship_df) if not ship_df.empty else pd.DataFrame()
    ws4 = wb.create_sheet("Сводная")
    summary_cols = ["Дата", "Регион", "Авто", "Рейс", "Тип тарифа", "Тариф",
                     "Кол-во заказов", "Кол-во магазинов", "Кол-во штук",
                     "Сумма итого", "Сумма распределенная"]
    ws4.append(summary_cols)
    for c in range(1, len(summary_cols) + 1):
        ws4.cell(row=1, column=c).font = HEADER_FONT

    n_summary_rows = 0
    if not summary.empty:
        for _, row in summary.iterrows():
            ws4.append(
                [
                    row["Дата"].date() if hasattr(row["Дата"], "date") else row["Дата"],
                    row["Регион"],
                    row["Авто"],
                    row["Рейс"],
                    row["Тип_тарифа"],
                    row["Тариф"],
                    row["Кол_во_заказов"],
                    row["Кол_во_магазинов"],
                    row["Кол_во_шт"],
                    row["Сумма_итого"],
                    row["Сумма_распределенная"],
                ]
            )
            n_summary_rows += 1
            if row["Тип_тарифа"] == "Сквозной":
                for c in range(1, len(summary_cols) + 1):
                    ws4.cell(row=ws4.max_row, column=c).fill = THROUGH_FILL

    for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row, min_col=1, max_col=1):
        row[0].number_format = DATE_FMT
    _autosize(ws4, len(summary_cols), max_width=40)

    if n_summary_rows > 0:
        last_col_letter = get_column_letter(len(summary_cols))
        table_ref = f"A1:{last_col_letter}{n_summary_rows + 1}"
        tbl2 = Table(displayName="ТаблицаСводная", ref=table_ref)
        tbl2.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False,
            showLastColumn=False, showColumnStripes=False,
        )
        ws4.add_table(tbl2)

    buf = _BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fix_date_columns(df: pd.DataFrame, date_cols) -> pd.DataFrame:
    """Приводит колонки с датами к чистому python date (без времени 00:00:00)."""
    df = df.copy()
    for col in date_cols:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col]).dt.date
    return df


def _apply_date_format(ws, col_letters, n_rows):
    for letter in col_letters:
        for r in range(2, n_rows + 2):
            ws[f"{letter}{r}"].number_format = "YYYY-MM-DD"


def build_excel_report(daily: pd.DataFrame, store_detail: pd.DataFrame, summary: pd.DataFrame,
                        registry: pd.DataFrame) -> bytes:
    """
    Собирает расширенный Excel-отчёт (без SLA — для него отдельный экспорт,
    см. build_sla_workbook) с листами:
      - "Отгрузки по дням"
      - "Детали по магазинам" (с колонкой "Город")
      - "Сводная"
      - "Реестр (аудит)" — построчный аудит исходных данных, из которых
        получены все агрегаты (по нему можно проверить/пересчитать любую
        цифру в отчёте вручную).
    """
    from io import BytesIO as _BytesIO

    daily = _fix_date_columns(daily, ["Дата"]) if daily is not None else daily
    store_detail = _fix_date_columns(store_detail, ["Дата"]) if store_detail is not None else store_detail
    registry = _fix_date_columns(registry, ["Дата"]) if registry is not None else registry

    buf = _BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        if daily is not None and not daily.empty:
            daily.to_excel(writer, sheet_name="Отгрузки по дням", index=False)
            _apply_date_format(writer.sheets["Отгрузки по дням"], ["A"], len(daily))
        if store_detail is not None and not store_detail.empty:
            store_detail.to_excel(writer, sheet_name="Детали по магазинам", index=False)
            _apply_date_format(writer.sheets["Детали по магазинам"], ["A"], len(store_detail))
        if summary is not None and not summary.empty:
            summary.to_excel(writer, sheet_name="Сводная", index=False)
        if registry is not None and not registry.empty:
            registry.to_excel(writer, sheet_name="Реестр (аудит)", index=False)
            _apply_date_format(writer.sheets["Реестр (аудит)"], ["A"], len(registry))

    return buf.getvalue()


def build_sla_workbook(sla: pd.DataFrame) -> bytes:
    """Отдельный Excel-файл с SLA (план vs факт), с чистыми датами (без времени)."""
    from io import BytesIO as _BytesIO

    sla = _fix_date_columns(sla, ["Дата_план", "Дата_факт"]) if sla is not None else sla

    buf = _BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        if sla is not None and not sla.empty:
            sla.to_excel(writer, sheet_name="SLA план-факт", index=False)
            _apply_date_format(writer.sheets["SLA план-факт"], ["D", "F"], len(sla))
        else:
            pd.DataFrame({"Инфо": ["Нет данных для SLA"]}).to_excel(writer, sheet_name="SLA план-факт", index=False)

    return buf.getvalue()
