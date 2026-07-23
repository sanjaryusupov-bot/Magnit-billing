"""
Парсинг листов гугл-таблицы MCOS в плоские (tidy) таблицы для дашборда.

Лист "Статус по отгрузкам" — это визуальный отчёт с группировками:
    День (строка с датой, без заказа)
      Рейс N (строка-заголовок)
        строки заказов (дата, заказ, кол-во шт, магазин, регион, авто, сумма...)

Лист "Статус по заказам" — тоже с группировками по дням в колонке A:
    "01/07/2026 (заказы на ср)"  <- заголовок раздела (это и есть план. дата)
      строки заказов (№ заказа, ID магазина, магазин, кол-во шт, дата отгрузки план, город, статусы...)
"""
import re
from datetime import datetime, date, timedelta
from io import BytesIO

import openpyxl
import pandas as pd

SHIPMENTS_SHEET = "Статус по отгрузкам"
ORDERS_SHEET = "Статус по заказам"
TARIFFS_SHEET = "тарифы"

_DATE_RE = re.compile(r"(\d{2})/(\d{2})/(\d{4})")

# Колонки (1-based), которые в исходных листах содержат дату и должны быть
# приведены к datetime, даже когда данные пришли как "сырые" числа (Google
# Sheets API отдаёт даты как порядковый номер дня, как в Excel).
_SHIPMENTS_DATE_COLS = {2}   # B: Дата отгрузки
_ORDERS_DATE_COLS = {6}      # F: Дата отгрузки план

_EXCEL_EPOCH = datetime(1899, 12, 30)

# Регион (точная точка доставки) -> Регион тариф (укрупнённая тарифная зона).
# По умолчанию — маппинг, извлечённый из реального файла биллинга заказчика
# (спутниковые города/районы тарифицируются по ближайшему хабу). Значения,
# которых нет в словаре, считаются тарифной зоной "как есть" (сами собой).
DEFAULT_REGION_TARIFF_MAP = {
    "Ангрен": "Ташкент",
    "Сырдарья": "Ташкент",
    "Ташобласть - Нурафшон": "Ташкент",
    "Ташобласть - Чирчик": "Ташкент",
    "Янгиюль": "Ташкент",
    "Коканд": "Фергана",
}

# Регионы, для которых при кол-ве штук >= 3000 на маршрут автоматически
# доплачивается "грузчик-экспедитор" (см. add_route_economics).
LONG_HAUL_LOADER_FEE_REGIONS = {"Самарканд", "Фергана", "Наманган", "Андижан", "Навои", "Бухара"}
LOADER_FEE_AMOUNT = 440000

# Справочник тарифов по умолчанию (копия листа "тарифы" из примера биллинга
# заказчика). Используется, если в самой Google-таблице нет листа "тарифы" —
# тогда экспорт всё равно будет содержать актуальный на момент настройки
# справочник. Если тарифы поменяются, обновите эту таблицу или добавьте лист
# "тарифы" в саму Google-таблицу — он будет использован автоматически.
DEFAULT_TARIFFS = [
    (1, "Ташкент", "Лабо с грузоподъемностью до 500кг. (9:00 - 18:00)",
     "Доставка заказов по г. Ташкент овтомобилем Лабо с грузоподъемностью до 500кг. (9:00 - 18:00)", 400000, 55000),
    (2, "Ташкент", "Лабо", "Доставка заказов по городу Ташкент - Лабо", 210000, 55000),
    (3, "Ташкент", "Газель", "Доставка заказов по городу Ташкент - Газель", 450000, 165000),
    (4, "Ташкент", "Исузу 5 тонн", "Доставка заказов по городу Ташкент - Исузу 5тонн", 935000, 165000),
    (5, "Ташкент", "10 тонн", "Доставка заказов по городу - 10 тонн", 1430000, 165000),
    (6, "Ташкент", "Евро фура", "Доставка заказов по городу - Евро фура", 1760000, 220000),
    (7, "Самарканд", "Газель", "Ташкент-Самарканд - Газель", 1900000, 165000),
    (8, "Самарканд", "Исузу 5 тонн", "Ташкент-Самарканд - Исузу 5тонн", 2420000, 165000),
    (9, "Самарканд", "10 тонн", "Ташкен-Самарканд - 10 тонн", 2970000, 220000),
    (10, "Фергана", "Газель", "Ташкент-Фергана - Газель", 2300000, 165000),
    (11, "Фергана", "Исузу 5 тонн", "Ташкент-Фергана - Исузу 5тонн", 2900000, 165000),
    (12, "Фергана", "10 тонн", "Ташкент-Фергана - 10 тонн", 3900000, 220000),
    (13, "Наманган", "Исузу 5 тонн", "Ташкент-Наманган - Исузу 5тонн", 2970000, 165000),
    (14, "Наманган", "10 тонн", "Ташкент-Наманган - 10 тонн", 3960000, 220000),
    (15, "Андижан", "Газель", "Ташкент-Андижан - Газель", 2600000, 165000),
    (16, "Андижан", "Исузу 5 тонн", "Ташкент-Андижан - Исузу 5тонн", 3300000, 220000),
    (17, "Навои", "2 тонны", "Ташкент-Навои 2 тонны", 2700000, 220000),
    (18, "Навои", "Исузу 5 тонн", "Ташкент-Навои 5 тонн", 4400000, 220000),
    (19, "Навои", "10 тонн", "Ташкент-Навои 10 тонн", 5900000, 220000),
    (20, "Бухара", "2 тонны", "Ташкент-Бухара 2 тонны", 3950000, 220000),
    (21, "Бухара", "Исузу 5 тонн", "Ташкент-Бухара 5 тонн", 4650000, 220000),
    (22, "Бухара", "10 тонн", "Ташкент-Бухара 10 тонн", 6500000, 220000),
    (23, None, None, "Дополнительно грузчик-экспедитор в машину (9:00-18:00)", 440000, None),
]

_TARIFF_COLS = ["№ п/п", "Регион", "Авто", "Маршрут", "Базовая ставка за рейс", "Стоимость дополнительной точки"]


def default_tariffs_df() -> pd.DataFrame:
    return pd.DataFrame(DEFAULT_TARIFFS, columns=_TARIFF_COLS)


def _serial_to_datetime(v):
    """Преобразует порядковый номер даты (Google Sheets/Excel) в datetime."""
    try:
        return _EXCEL_EPOCH + timedelta(days=float(v))
    except (TypeError, ValueError):
        return None


class _Cell:
    __slots__ = ("value",)

    def __init__(self, value):
        self.value = value


class ListWorksheet:
    """
    Адаптер: превращает список списков (сырые значения, например от gspread)
    в объект с интерфейсом openpyxl Worksheet (.cell(row, column).value,
    .max_row), чтобы parse_shipments/parse_orders_status работали одинаково
    и с .xlsx, и с данными из Google Sheets API.
    """

    def __init__(self, rows, date_columns=None):
        self.rows = rows
        self.date_columns = date_columns or set()
        self.max_row = len(rows)

    def cell(self, row, column):
        r, c = row - 1, column - 1
        if r < 0 or r >= len(self.rows):
            return _Cell(None)
        row_data = self.rows[r]
        if c < 0 or c >= len(row_data):
            return _Cell(None)
        v = row_data[c]
        if v == "":
            v = None
        if v is not None and column in self.date_columns and isinstance(v, (int, float)) and not isinstance(v, bool):
            dt = _serial_to_datetime(v)
            if dt is not None:
                v = dt
        return _Cell(v)


def _is_error(v):
    if v is None:
        return True
    if isinstance(v, str) and v.strip().startswith("#"):
        return True
    return False


def _num(v):
    if _is_error(v):
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _clean_str(v):
    if _is_error(v):
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def parse_shipments(ws) -> pd.DataFrame:
    """Лист 'Статус по отгрузкам' -> tidy DataFrame, 1 строка = 1 заказ в рейсе."""
    rows = []
    current_day = None
    current_reys = None
    for r in range(2, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value  # Дата отгрузки
        c = ws.cell(row=r, column=3).value  # Заказ
        if b is None and c is None:
            continue
        if isinstance(b, str) and b.strip().lower().startswith("рейс"):
            current_reys = b.strip()
            continue
        if isinstance(b, datetime) and c is None:
            current_day = b.date()
            current_reys = None
            continue
        if c is not None:
            day_val = b.date() if isinstance(b, datetime) else current_day
            rows.append(
                {
                    "Дата": day_val,
                    "Рейс": current_reys,
                    "Заказ": str(c).strip(),
                    "Кол_во_шт": _num(ws.cell(row=r, column=4).value),
                    "ID_магазина": _clean_str(ws.cell(row=r, column=5).value),
                    "Магазин": _clean_str(ws.cell(row=r, column=6).value),
                    "Адрес": _clean_str(ws.cell(row=r, column=7).value),
                    "Координаты_1": _clean_str(ws.cell(row=r, column=8).value),
                    "Координаты_2": _clean_str(ws.cell(row=r, column=9).value),
                    "Регион": _clean_str(ws.cell(row=r, column=10).value),
                    "Транспорт": _clean_str(ws.cell(row=r, column=11).value),
                    "Точка_1": _num(ws.cell(row=r, column=12).value),
                    "Точка_2_и_далее": _num(ws.cell(row=r, column=13).value),
                    "Грузчик_экспедитор": _num(ws.cell(row=r, column=14).value),
                    "Итого_сумма": _num(ws.cell(row=r, column=15).value),
                    "Подрядчик": _clean_str(ws.cell(row=r, column=16).value),
                }
            )
    df = pd.DataFrame(rows)
    if not df.empty:
        df["Дата"] = pd.to_datetime(df["Дата"])
    return df


def parse_orders_status(ws) -> pd.DataFrame:
    """Лист 'Статус по заказам' -> tidy DataFrame с плановой датой отгрузки."""
    rows = []
    current_section_date = None
    for r in range(2, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if a is None:
            continue
        a_str = str(a).strip()
        if not a_str.startswith("1M-"):
            m = _DATE_RE.search(a_str)
            if m:
                d, mo, y = m.groups()
                try:
                    current_section_date = date(int(y), int(mo), int(d))
                except ValueError:
                    pass
            continue
        plan_cell = ws.cell(row=r, column=6).value  # Дата отгрузки план
        plan_date = plan_cell.date() if isinstance(plan_cell, datetime) else current_section_date
        rows.append(
            {
                "Заказ": a_str,
                "ID_магазина": _clean_str(ws.cell(row=r, column=2).value),
                "Магазин": _clean_str(ws.cell(row=r, column=3).value),
                "Адрес_магазина": _clean_str(ws.cell(row=r, column=4).value),
                "Кол_во_шт_заказ": _num(ws.cell(row=r, column=5).value),
                "Дата_план": plan_date,
                "Город": _clean_str(ws.cell(row=r, column=7).value),
                "Статус_сборки": _clean_str(ws.cell(row=r, column=8).value),
                "Статус_WMS": _clean_str(ws.cell(row=r, column=9).value),
                "Статус_отгрузки_на_хаб": _clean_str(ws.cell(row=r, column=10).value),
            }
        )
    df = pd.DataFrame(rows)
    if not df.empty:
        df["Дата_план"] = pd.to_datetime(df["Дата_план"])
    return df


def add_route_economics(ship_df: pd.DataFrame, region_tariff_map: dict = None) -> pd.DataFrame:
    """
    Пересчитывает стоимость доставки по логике заказчика и добавляет поля,
    нужные для биллинг-отчёта.

    В исходнике "Итого сумма" — это тариф за точку маршрута (магазин), который
    проставлен только один раз на уникальный магазин в рейсе (повторные заказы
    в тот же магазин в тот же рейс получают 0, доставка уже учтена). Из-за
    этого при группировке "по магазину" суммы получаются неровными — то 0, то
    полный тариф.

    Правильный расчёт:
        1. Сумма_маршрута = сумма "Итого сумма" по всем строкам рейса
           (это и есть полная стоимость доставки рейса — сумма тарифов по
           каждой уникальной точке, т.к. повторы уже дают 0).
        2. Кол_во_шт_маршрута = сумма "Кол-во шт" по всем строкам рейса
           (включая повторные заказы в тот же магазин — они тоже везли товар).
        3. Тариф_за_шт = Сумма_маршрута / Кол_во_шт_маршрута.
        4. Для каждой строки (заказа): Сумма_распределённая = Кол-во_шт
           этой строки * Тариф_за_шт — то есть стоимость доставки размазана
           по всем штукам маршрута пропорционально, а не только на первую
           точку.

    Дополнительно (для биллинг-реестра, формат как в примере заказчика) —
    ДВА варианта идентификатора маршрута:
        - Маршрут_ID_первый_заказ — простой вариант: номер ПЕРВОГО заказа,
          встреченного в блоке (Дата, Рейс) исходника, без какого-либо
          разделения (как в самом первом варианте реестра). Идёт в колонку
          "Рейс" биллинг-отчёта.
        - Маршрут_ID — "умный" вариант с разделением ПО РЕГИОНУ: в пределах
          одного "Рейс N" все строки с одинаковым Регион_тариф считаются
          одним маршрутом (Маршрут_ID = номер первого заказа этого региона
          в блоке), а строки с другим регионом — другим маршрутом, даже
          если регионы в блоке чередуются (Наманган, Коканд, Наманган ->
          ровно 2 маршрута: один на все строки Наманган, другой на все
          строки Коканд). Идёт в колонку "Рейс 2" биллинг-отчёта и
          используется для расчёта Тариф_за_шт/Сумма_распределенная (это
          финансово корректная группировка).
        - Регион_тариф — регион, нормализованный к тарифной зоне (например,
          Янгиюль/Ангрен/Чирчик -> Ташкент, Коканд -> Фергана), по словарю
          region_tariff_map (по умолчанию DEFAULT_REGION_TARIFF_MAP).

    Доплата "грузчик-экспедитор" (LONG_HAUL_LOADER_FEE_REGIONS): для
    маршрутов (Дата, Маршрут_ID) с регионом из списка Самарканд, Фергана,
    Наманган, Андижан, Навои, Бухара (Коканд считается Фергана) — если
    суммарное кол-во штук по маршруту >= 3000, к строке-якорю маршрута
    добавляется 440 000 в "Грузчик экспедитор" и в "Итого сумма" (если там
    уже не было учтено). Для остальных регионов (в т.ч. Ташкент) поведение
    не меняется.

    Группировка для расчёта тарифа за штуку — по (Дата, Маршрут_ID), т.к.
    именно это и есть настоящий физический маршрут (в отличие от "Рейс N",
    который может объединять несколько разных маршрутов с разными тарифами).
    """
    if ship_df.empty:
        return ship_df

    region_map = region_tariff_map if region_tariff_map is not None else DEFAULT_REGION_TARIFF_MAP

    df = ship_df.copy()
    df["Регион_тариф"] = df["Регион"].map(lambda r: region_map.get(r, r) if r is not None else r)

    # Маршрут_ID_первый_заказ: наивный вариант — просто номер первого заказа
    # в блоке (Дата, Рейс), без разделения. Идёт в колонку "Рейс".
    df["Маршрут_ID_первый_заказ"] = df.groupby(
        ["Дата", "Рейс"], dropna=False, sort=False
    )["Заказ"].transform("first")

    # Маршрут_ID: в пределах одного блока (Дата, Рейс) группируем СТРОГО по
    # Регион_тариф — все строки одного региона получают номер первого
    # заказа этого региона в блоке, независимо от того, чередуются ли
    # регионы построчно (Наманган/Коканд/Наманган -> 2 маршрута, не 3).
    def _assign_route_ids(g: pd.DataFrame) -> pd.Series:
        orders = g["Заказ"].tolist()
        regions = g["Регион_тариф"].tolist()
        region_to_anchor = {}
        result = []
        for order, region in zip(orders, regions):
            anchor = region_to_anchor.setdefault(region, order)
            result.append(anchor)
        return pd.Series(result, index=g.index)

    marshrut_id = pd.Series(index=df.index, dtype=object)
    for _, idx in df.groupby(["Дата", "Рейс"], dropna=False, sort=False).groups.items():
        marshrut_id.loc[idx] = _assign_route_ids(df.loc[idx])
    df["Маршрут_ID"] = marshrut_id

    # Доплата "грузчик-экспедитор" для междугородних маршрутов большого
    # объёма (см. докстринг). Считаем кол-во штук по (Дата, Маршрут_ID) ДО
    # добавления доплаты (сама доплата не зависит от кол-ва штук).
    qty_per_route = df.groupby(["Дата", "Маршрут_ID"], dropna=False)["Кол_во_шт"].transform("sum")
    region_per_route = df.groupby(["Дата", "Маршрут_ID"], dropna=False)["Регион_тариф"].transform("first")
    is_anchor_row = df["Заказ"] == df["Маршрут_ID"]
    eligible = (
        region_per_route.isin(LONG_HAUL_LOADER_FEE_REGIONS)
        & (qty_per_route >= 3000)
        & is_anchor_row
    )
    df.loc[eligible, "Грузчик_экспедитор"] = df.loc[eligible, "Грузчик_экспедитор"] + LOADER_FEE_AMOUNT
    df.loc[eligible, "Итого_сумма"] = df.loc[eligible, "Итого_сумма"] + LOADER_FEE_AMOUNT

    grp_cols = ["Дата", "Маршрут_ID"]
    route_totals = (
        df.groupby(grp_cols, dropna=False)
        .agg(Сумма_маршрута=("Итого_сумма", "sum"), Кол_во_шт_маршрута=("Кол_во_шт", "sum"))
        .reset_index()
    )
    route_totals["Тариф_за_шт"] = route_totals.apply(
        lambda r: (r["Сумма_маршрута"] / r["Кол_во_шт_маршрута"]) if r["Кол_во_шт_маршрута"] else 0.0,
        axis=1,
    )
    df = df.merge(route_totals, on=grp_cols, how="left")
    df["Сумма_распределенная"] = df["Кол_во_шт"] * df["Тариф_за_шт"]

    return df


def parse_tariffs(ws) -> pd.DataFrame:
    """Лист 'тарифы' (справочник ставок) -> DataFrame с теми же заголовками."""
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 7)]
        if all(v is None for v in vals):
            continue
        rows.append(vals)
    if not rows:
        return pd.DataFrame(columns=_TARIFF_COLS)
    return pd.DataFrame(rows, columns=_TARIFF_COLS)


def parse_workbook_bytes(data: bytes):
    """Возвращает (ship_df, orders_df) из байтов .xlsx файла."""
    wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
    ship_df = parse_shipments(wb[SHIPMENTS_SHEET])
    orders_df = parse_orders_status(wb[ORDERS_SHEET])
    return ship_df, orders_df


_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets.readonly",
    "https://www.googleapis.com/auth/drive.readonly",
]


def load_from_gsheet_service_account(sheet_id: str, credentials_info: dict):
    """
    Возвращает (ship_df, orders_df, tariffs_df), читая таблицу через Google
    Sheets API с помощью сервисного аккаунта. Таблицу не нужно открывать по
    ссылке — достаточно выдать доступ на чтение самому сервисному аккаунту
    (credentials_info['client_email']).

    Лист "тарифы" — необязательный: если такого листа в таблице нет,
    возвращается пустой DataFrame (в приложении в этом случае используется
    встроенный справочник по умолчанию, см. default_tariffs_df()).
    """
    import gspread
    from google.oauth2.service_account import Credentials

    creds = Credentials.from_service_account_info(dict(credentials_info), scopes=_SCOPES)
    client = gspread.authorize(creds)
    sh = client.open_by_key(sheet_id)

    ship_ws = sh.worksheet(SHIPMENTS_SHEET)
    orders_ws = sh.worksheet(ORDERS_SHEET)

    ship_rows = ship_ws.get_values(value_render_option="UNFORMATTED_VALUE")
    orders_rows = orders_ws.get_values(value_render_option="UNFORMATTED_VALUE")

    ship_adapter = ListWorksheet(ship_rows, date_columns=_SHIPMENTS_DATE_COLS)
    orders_adapter = ListWorksheet(orders_rows, date_columns=_ORDERS_DATE_COLS)

    ship_df = parse_shipments(ship_adapter)
    orders_df = parse_orders_status(orders_adapter)

    tariffs_df = pd.DataFrame(columns=_TARIFF_COLS)
    try:
        tariffs_ws = sh.worksheet(TARIFFS_SHEET)
        tariffs_rows = tariffs_ws.get_values(value_render_option="UNFORMATTED_VALUE")
        tariffs_adapter = ListWorksheet(tariffs_rows)
        tariffs_df = parse_tariffs(tariffs_adapter)
    except Exception:
        pass  # листа "тарифы" нет в таблице — это нормально, используем дефолт

    return ship_df, orders_df, tariffs_df


def build_sla(ship_df: pd.DataFrame, orders_df: pd.DataFrame) -> pd.DataFrame:
    """Джойн план (Статус по заказам) vs факт (Статус по отгрузкам) по номеру заказа 1M-."""
    if ship_df.empty or orders_df.empty:
        return pd.DataFrame()

    fact = (
        ship_df[ship_df["Заказ"].str.startswith("1M-")]
        .groupby("Заказ", as_index=False)
        .agg(Дата_факт=("Дата", "min"))
    )
    plan = orders_df[orders_df["Заказ"].str.startswith("1M-")][
        ["Заказ", "Магазин", "Город", "Дата_план", "Статус_отгрузки_на_хаб"]
    ]
    sla = plan.merge(fact, on="Заказ", how="left")
    sla["Дельта_дней"] = (sla["Дата_факт"] - sla["Дата_план"]).dt.days
    sla["SLA_статус"] = sla.apply(_sla_flag, axis=1)
    return sla


def _sla_flag(row):
    if pd.isna(row["Дата_план"]):
        return "Нет плана"
    if pd.isna(row["Дата_факт"]):
        return "Не отгружен"
    if row["Дельта_дней"] <= 0:
        return "В срок"
    return "Просрочка"


def build_billing_workbook(ship_df: pd.DataFrame) -> bytes:
    """
    Собирает Excel-файл с одним листом "Реестр" — построчно по каждому
    заказу, колонки как в примере заказчика (Дата отгрузки, Заказ, Рейс,
    Кол-во штук, ID магазина, Магазин, Адрес, Координаты x2, Регион,
    Регион тариф, Авто, 1 точка, 2 точка и далее, Грузчик экспедитор,
    Итого сумма). Оформлен как настоящая Excel-таблица (с автофильтром) —
    источник данных, из которого в Excel за 2 клика строится живая сводная
    таблица с двойным кликом "Показать детали" (см. README).

    ship_df должен быть уже отфильтрован (период, только 1M- и т.п.) и
    содержать колонки, которые добавляет add_route_economics().
    """
    from io import BytesIO as _BytesIO

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = Workbook()
    DATE_FMT = "YYYY-MM-DD"

    def _autosize(ws, ncols, min_width=10, max_width=45):
        for c in range(1, ncols + 1):
            letter = get_column_letter(c)
            max_len = min_width
            for row in ws.iter_rows(min_col=c, max_col=c):
                v = row[0].value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[letter].width = min(max_len + 2, max_width)

    ws3 = wb.active
    ws3.title = "Реестр"
    registry_cols = [
        "Дата отгрузки", "Заказ", "Рейс", "Рейс 2", "Кол-во штук", "ID магазина", "Магазин", "Адрес",
        "Координаты", "Координаты2", "Регион", "Регион тариф", "Авто",
        "1 точка", "2 точка и далее", "Грузчик экспедитор", "Итого сумма",
    ]
    ws3.append(registry_cols)

    n_data_rows = 0
    if not ship_df.empty:
        reg = ship_df.sort_values(["Дата", "Маршрут_ID", "Заказ"])
        for _, row in reg.iterrows():
            ws3.append(
                [
                    row["Дата"].date(),
                    row["Заказ"],
                    row["Маршрут_ID_первый_заказ"],
                    row["Маршрут_ID"],
                    row["Кол_во_шт"],
                    row["ID_магазина"],
                    row["Магазин"],
                    row["Адрес"],
                    row["Координаты_1"],
                    row["Координаты_2"],
                    row["Регион"],
                    row["Регион_тариф"],
                    row["Транспорт"],
                    row["Точка_1"],
                    row["Точка_2_и_далее"],
                    row["Грузчик_экспедитор"],
                    row["Итого_сумма"],
                ]
            )
            n_data_rows += 1

    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=1, max_col=1):
        row[0].number_format = DATE_FMT
    _autosize(ws3, len(registry_cols), max_width=40)

    if n_data_rows > 0:
        last_col_letter = get_column_letter(len(registry_cols))
        table_ref = f"A1:{last_col_letter}{n_data_rows + 1}"
        tbl = Table(displayName="ТаблицаРеестр", ref=table_ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False,
            showLastColumn=False, showColumnStripes=False,
        )
        ws3.add_table(tbl)

    buf = _BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fix_date_columns(df: pd.DataFrame, date_cols) -> pd.DataFrame:
    """Приводит колонки с датами к чистому python date (без времени 00:00:00)."""
    df = df.copy()
    for col in date_cols:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col]).dt.date
    return df


def _apply_date_format(ws, col_letters, n_rows):
    for letter in col_letters:
        for r in range(2, n_rows + 2):
            ws[f"{letter}{r}"].number_format = "YYYY-MM-DD"


def build_excel_report(daily: pd.DataFrame, store_detail: pd.DataFrame, summary: pd.DataFrame,
                        registry: pd.DataFrame) -> bytes:
    """
    Собирает расширенный Excel-отчёт (без SLA — для него отдельный экспорт,
    см. build_sla_workbook) с листами:
      - "Отгрузки по дням"
      - "Детали по магазинам" (с колонкой "Город")
      - "Сводная"
      - "Реестр (аудит)" — построчный аудит исходных данных, из которых
        получены все агрегаты (по нему можно проверить/пересчитать любую
        цифру в отчёте вручную).
    """
    from io import BytesIO as _BytesIO

    daily = _fix_date_columns(daily, ["Дата"]) if daily is not None else daily
    store_detail = _fix_date_columns(store_detail, ["Дата"]) if store_detail is not None else store_detail
    registry = _fix_date_columns(registry, ["Дата"]) if registry is not None else registry

    buf = _BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        if daily is not None and not daily.empty:
            daily.to_excel(writer, sheet_name="Отгрузки по дням", index=False)
            _apply_date_format(writer.sheets["Отгрузки по дням"], ["A"], len(daily))
        if store_detail is not None and not store_detail.empty:
            store_detail.to_excel(writer, sheet_name="Детали по магазинам", index=False)
            _apply_date_format(writer.sheets["Детали по магазинам"], ["A"], len(store_detail))
        if summary is not None and not summary.empty:
            summary.to_excel(writer, sheet_name="Сводная", index=False)
        if registry is not None and not registry.empty:
            registry.to_excel(writer, sheet_name="Реестр (аудит)", index=False)
            _apply_date_format(writer.sheets["Реестр (аудит)"], ["A"], len(registry))

    return buf.getvalue()


def build_sla_workbook(sla: pd.DataFrame) -> bytes:
    """Отдельный Excel-файл с SLA (план vs факт), с чистыми датами (без времени)."""
    from io import BytesIO as _BytesIO

    sla = _fix_date_columns(sla, ["Дата_план", "Дата_факт"]) if sla is not None else sla

    buf = _BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        if sla is not None and not sla.empty:
            sla.to_excel(writer, sheet_name="SLA план-факт", index=False)
            _apply_date_format(writer.sheets["SLA план-факт"], ["D", "F"], len(sla))
        else:
            pd.DataFrame({"Инфо": ["Нет данных для SLA"]}).to_excel(writer, sheet_name="SLA план-факт", index=False)

    return buf.getvalue()
